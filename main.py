from datetime import datetime, date, timedelta
from pathlib import Path
import streamlit as st
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import requests
import streamlit as st
from streamlit_lottie import st_lottie
import pandas as pd
from io import StringIO
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
import base64
from PIL import Image
import io
import os

# Page configuration
st.set_page_config(
    page_title="EU Taxonomy",
    page_icon="🌍",
    layout="wide",
    initial_sidebar_state="expanded"
)

file_path = 'Project Damietta_CashFlow Model_01b.xlsx'

os.chmod(file_path, 0o666)

@st.cache_data
def load_financial_model(file_path,sheet_name,header):
    file_extension = Path(file_path).suffix.lower()[1:]

    if file_extension == 'xlsx':
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header, engine='openpyxl')

    elif file_extension == 'xls':
        st.write('its a xls file')
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header,engine='openpyxl')
    elif file_extension == 'csv':
        return pd.read_csv(file_path)
    else:
        raise Exception("File not supported")

# Define a common background color
bg_color = "#000C66"  # Adjust this color as needed
check_phase1 = False
check_phase2 = False

# Initialize session state variables at the start
if 'page' not in st.session_state:
    st.session_state.page = 'main'

if 'show_eligibility' not in st.session_state:
    st.session_state.show_eligibility = False

if 'selected_phase_option' not in st.session_state:
    st.session_state.selected_phase_option = "Select an option"

# Initialize all form field states if not already present
def init_session_state():
    default_fields = {
        # Phase 1 fields
        'fcp1': datetime.now().date(),  # Set default date to today
        'field7': datetime.now().date(),  # Set default date to today
        'field8': 0.0,
        'field9': None,
        'field10': 0.0,
        'field11': 0.0,
        'field12': 0.0,
        'upfp1': 0.0,
        'cfp1': 0.0,
        'field14': 0.0,
        'field15': 0.0,
        'field16': 0.0,
        'field17': 0.0,
        'field18': 0.0,
        'field19': 0.0,
        'ofwaccp1': 0.0,
        'drcitrp1': 0.0,
        'cgp1': 0.0,
        'cfltp1': 0.0,
        'ogsgp1': 0.0,
        'ofo1': 0.0,
        'field25': 0.0,
        'field26': 0.0,
        'field27': 0.0,
        'field28': 0.0,
        'field29': 0.0,
        'field34': 0.0,
        'drtp1': 0.0,
        'cepsp1': 0.0,
        'ces1': 0.0,
        
        # Phase 2 fields
        'fcp2': datetime.now().date(),  # Set default date to today
        'field7p2': datetime.now().date(),  # Set default date to today
        'field8p2': 0.0,
        'field9p2': None,
        'field10p2': 0.0,
        'field11p2': 0.0,
        'field12p2': 0.0,
        'upfp2': 0.0,
        'cfp2': 0.0,
        'field14p2': 0.0,
        'field15p2': 0.0,
        'field16p2': 0.0,
        'field17p2': 0.0,
        'field18p2': 0.0,
        'field19p2': 0.0,
        'ofwaccp2': 0.0,
        'drcitrp2': 0.0,
        'cgp2': 0.0,
        'cfltp2': 0.0,
        'ogsgp2': 0.0,
        'ofo2': 0.0,
        'field25p2': 0.0,
        'field26p2': 0.0,
        'field27p2': 0.0,
        'field28p2': 0.0,
        'field29p2': 0.0,
        'field34p2': 0.0,
        'drtp2': 0.0,
        'cepsp2': 0.0,
        'ces2': 0.0,
        
        # Other state variables
        'answer7': 0.0,
        'answer8': 0.0,
        'osdp1': None,
        'oedp1': None,
        'drsdp1': None,
        'dredp1': None,
        'osdp2': None,
        'oedp2': None,
        'drsdp2': None,
        'dredp2': None
    }
    
    for field, default_value in default_fields.items():
        if field not in st.session_state:
            st.session_state[field] = default_value

# Call initialization at the start
init_session_state()

# Update the answer7 input to use session state
def handle_answer7_change():
    st.session_state.answer7 = st.session_state.get('answer7_input', 0.0)

# Update the answer8 input to use session state
def handle_answer8_change():
    st.session_state.answer8 = st.session_state.get('answer8_input', 0.0)

# Main page form fields
if 'field1' not in st.session_state:
    st.session_state.field1 = ""
# Phase 1 form fields
if 'fcp1' not in st.session_state:
    st.session_state.fcp1 = None
if 'field7' not in st.session_state:
    st.session_state.field7 = None
if 'field8' not in st.session_state:
    st.session_state.field8 = 0.0
if 'field9' not in st.session_state:
    st.session_state.field9 = None
if 'field10' not in st.session_state:
    st.session_state.field10 = 0.0
if 'field11' not in st.session_state:
    st.session_state.field11 = 0.0
if 'field12' not in st.session_state:
    st.session_state.field12 = 0.0
if 'field14' not in st.session_state:
    st.session_state.field14 = 0.0
if 'field15' not in st.session_state:
    st.session_state.field15 = 0.0
if 'field16' not in st.session_state:
    st.session_state.field16 = 0.0
if 'field17' not in st.session_state:
    st.session_state.field17 = 0.0
if 'field18' not in st.session_state:
    st.session_state.field18 = 0.0
if 'field19' not in st.session_state:
    st.session_state.field19 = 0.0
if 'field25' not in st.session_state:
    st.session_state.field25 = 0.0
if 'field26' not in st.session_state:
    st.session_state.field26 = 0.0
if 'field27' not in st.session_state:
    st.session_state.field27 = 0.0
if 'field28' not in st.session_state:
    st.session_state.field28 = 0.0
if 'field29' not in st.session_state:
    st.session_state.field29 = 0.0
if 'field34' not in st.session_state:
    st.session_state.field34 = 0.0
if 'osdp1' not in st.session_state:
    st.session_state.osdp1 = None
if 'drsdp1' not in st.session_state:
    st.session_state.drsdp1 = None
if 'drtp1' not in st.session_state:
    st.session_state.drtp1 = 0.0
if 'dredp1' not in st.session_state:
    st.session_state.dredp1 = None
if 'cepsp1' not in st.session_state:
    st.session_state.cepsp1 = 0.0
if 'ces1' not in st.session_state:
    st.session_state.ces1 = 0.0
if 'upfp1' not in st.session_state:
    st.session_state.upfp1 = 0.0
if 'cfp1' not in st.session_state:
    st.session_state.cfp1 = 0.0
if 'cgp1' not in st.session_state:
    st.session_state.cgp1 = 0.0
if 'cfltp1' not in st.session_state:
    st.session_state.cfltp1 = 0.0
if 'ogsgp1' not in st.session_state:
    st.session_state.ogsgp1 = 0.0
if 'ofo1' not in st.session_state:
    st.session_state.ofo1 = 0.0
if 'ofwaccp1' not in st.session_state:
    st.session_state.ofwaccp1 = 0.0
if 'drcitrp1' not in st.session_state:
    st.session_state.drcitrp1 = 0.0

# Phase 2 form fields
if 'fcp2' not in st.session_state:
    st.session_state.fcp2 = None
if 'field35' not in st.session_state:
    st.session_state.field35 = None
if 'field36' not in st.session_state:
    st.session_state.field36 = 0.0
if 'field37' not in st.session_state:
    st.session_state.field37 = None
if 'field38' not in st.session_state:
    st.session_state.field38 = 0.0
if 'field40' not in st.session_state:
    st.session_state.field40 = 0.0
if 'field42' not in st.session_state:
    st.session_state.field42 = 0.0
if 'field43' not in st.session_state:
    st.session_state.field43 = 0.0
if 'field44' not in st.session_state:
    st.session_state.field44 = 0.0
if 'field45' not in st.session_state:
    st.session_state.field45 = 0.0
if 'field46' not in st.session_state:
    st.session_state.field46 = 0.0
if 'field47' not in st.session_state:
    st.session_state.field47 = 0.0
if 'field53' not in st.session_state:
    st.session_state.field53 = 0.0
if 'field54' not in st.session_state:
    st.session_state.field54 = 0.0
if 'field58' not in st.session_state:
    st.session_state.field58 = 0.0
if 'field59' not in st.session_state:
    st.session_state.field59 = 0.0
if 'field60' not in st.session_state:
    st.session_state.field60 = 0.0
if 'field62' not in st.session_state:
    st.session_state.field62 = 0.0
if 'field69' not in st.session_state:
    st.session_state.field69 = 0.0
if 'osdp2' not in st.session_state:
    st.session_state.osdp2 = None
if 'drsdp2' not in st.session_state:
    st.session_state.drsdp2 = None
if 'drtp2' not in st.session_state:
    st.session_state.drtp2 = 0.0
if 'dredp2' not in st.session_state:
    st.session_state.dredp2 = None
if 'cepsp2' not in st.session_state:
    st.session_state.cepsp2 = 0.0
if 'ces2' not in st.session_state:
    st.session_state.ces2 = 0.0
if 'upfp2' not in st.session_state:
    st.session_state.upfp2 = 0.0
if 'cfp2' not in st.session_state:
    st.session_state.cfp2 = 0.0
if 'cgp2' not in st.session_state:
    st.session_state.cgp2 = 0.0
if 'cfltp2' not in st.session_state:
    st.session_state.cfltp2 = 0.0
if 'ogsgp2' not in st.session_state:
    st.session_state.ogsgp2 = 0.0
if 'ofo2' not in st.session_state:
    st.session_state.ofo2 = 0.0
if 'ofwaccp2' not in st.session_state:
    st.session_state.ofwaccp2 = 0.0
if 'drcitrp2' not in st.session_state:
    st.session_state.drcitrp2 = 0.0

# Risk assessment data
if 'risk_assessment_df' not in st.session_state:
    st.session_state.risk_assessment_df = None
if 'selected_risks' not in st.session_state:
    st.session_state.selected_risks = ["Select a risk"] * 18

# Main page form fields
if 'field1' not in st.session_state:
    st.session_state.field1 = ""
if 'field2' not in st.session_state:
    st.session_state.field2 = ""
if 'field3' not in st.session_state:
    st.session_state.field3 = ""
if 'field4' not in st.session_state:
    st.session_state.field4 = ""
if 'field5' not in st.session_state:
    st.session_state.field5 = ""

# CSS style for larger font size
if 'field1' not in st.session_state:
    st.session_state.field1 = ""
if 'field2' not in st.session_state:
    st.session_state.field2 = ""
if 'field3' not in st.session_state:
    st.session_state.field3 = ""
if 'field4' not in st.session_state:
    st.session_state.field4 = ""
if 'field5' not in st.session_state:
    st.session_state.field5 = ""
if 'field9' not in st.session_state:
    st.session_state.field9 = ""
if 'field11' not in st.session_state:
    st.session_state.field11 = ""
if 'field16' not in st.session_state:
    st.session_state.field16 = 0.0
if 'field19' not in st.session_state:
    st.session_state.field19 = 0.0
if 'field27' not in st.session_state:
    st.session_state.field27 = 0.0
if 'field32' not in st.session_state:
    st.session_state.field32 = 0.0
if 'field34' not in st.session_state:
    st.session_state.field34 = 0.0
if 'field37' not in st.session_state:
    st.session_state.field37 = 0.0
if 'field39' not in st.session_state:
    st.session_state.field39 = 0.0
if 'field44' not in st.session_state:
    st.session_state.field44 = 0.0
if 'field47' not in st.session_state:
    st.session_state.field47 = 0.0
if 'field60' not in st.session_state:
    st.session_state.field60 = 0.0

if 'osdp1' not in st.session_state:
        st.session_state.osdp1 = 0.0

if 'field17' not in st.session_state:
        st.session_state.field17 = 0.0

if 'osdp2' not in st.session_state:
    st.session_state.osdp2 = None

st.markdown(
    """
    <style>
    .big-font-section {
        font-size: 18px !important;
        text-align: center;
    }
    .big-font {
        font-size: 18px !important;
    }
    .center-text {
        text-align: center;
        margin: 0 auto;
        width: 70%; /* Adjust the width as needed */
    }
    .info-button {
        margin-left: 10px;
        background-color: #007BFF;
        border: none;
        color: white;
        padding: 5px 10px;
        border-radius: 50%;
        cursor: pointer;
        display: inline-block;
        width: 30px; /* Adjust the width and height to ensure the button is circular */
        height: 30px; /* Adjust the width and height to ensure the button is circular */
        text-align: center;
        line-height: 20px; /* Align the text vertically in the center */
        float: right;
    }
    .tooltip {
        position: relative;
        display: inline-block;
        cursor: pointer;
        margin-left: 10px;
    }
    .tooltip .tooltiptext {
        visibility: hidden;
        width: 220px;
        background-color: #555;
        color: #fff;
        text-align: center;
        border-radius: 5px;
        padding: 5px 10px;
        position: absolute;
        z-index: 1;
        bottom: 150%; /* Adjust this value to move the tooltip */
        left: 50%;
        margin-left: -110px;
        opacity: 0;
        transition: opacity 0.3s;
        font-size: 15px !important;
    }
    .tooltip:hover .tooltiptext {
        visibility: visible;
        opacity: 1;
    }
    .bold-hr {
        border: 0;
        height: 5px; /* Adjust height to make it more bold */
        background: #000; /* Change color if needed */
    }
    </style>
    """,
    unsafe_allow_html=True,
)
if 'show_eligibility' not in st.session_state:
        st.session_state.show_eligibility = False

def go_back_to_main_page():
    st.session_state.page = 'main'
    st.session_state.show_eligibility = False

def go_back_to_phase():
    st.session_state.page = 'phase'

def continue_to_phase():
    st.session_state.page = 'phase'

def continue_to_phase2():
    st.session_state.page = 'phase2'

def continue_to_risk_management():
    st.session_state.page = 'risk-management'

def continue_to_dashboard():
    st.session_state.page = 'dashboard'
    
# Sidebar for input fields and logos
if st.session_state.page == 'main':


    with st.sidebar:
        # Logos at the top of the sidebar
        col1, col2 = st.columns((1, 1))
        with col1:
            st.image("egypt.jpg", width=100)
        with col2:
            st.image("eu.jpg", width=100)
        
        st.markdown("# User Details")
        # field1 = ""
        # field2 = ""
        # field3 = ""
        # field4 = ""
        # field5 = ""
        st.session_state.field1 = st.text_input("Username", key='1', placeholder="Enter username")
        st.session_state.field2 = st.text_input("Project", key='2', placeholder="Enter project name")
        st.session_state.field3 = st.text_input("Capacity", key='3', placeholder="Enter capacity in m3/d")
        st.session_state.field4 = st.text_input("Location", key='4', placeholder="Enter location")
        st.session_state.field5 = st.date_input("Date", key='5')

        if st.button('Next') and st.session_state.field1 != "" and st.session_state.field2 != "" and st.session_state.field3 != "" and st.session_state.field4 != "" and st.session_state.field5 != "":
            st.session_state['show_eligibility'] = True

        st.markdown("<br><br>", unsafe_allow_html=True)

        st.image("urbane.jpeg",width=150)


        st.markdown(
            """
            <div style='text-align: center; font-family: Arial, sans-serif; font-size: 14px; color: #4CAF50; margin-top: 15px;'>
                Developed by <b>GreenUrbane</b>
            </div>
            """,
        unsafe_allow_html=True
)

    # Main content
    with st.container():
        col1, col2 = st.columns((1, 5))
        with col1:
            st.image("CL3.png",width=180)
        with col2:
            st.markdown(
                f'<br><br><div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 15px; margin-bottom: 15px; width: 100%; text-align: center; font-size: 36px; margin-top: -50px;" class="center-text">'
                '<strong>EU-WATER-FIT (Water Assessment and EU Taxonomy Evaluation for Resilient Financing and Investment Tool)</strong>'
                '</div>', unsafe_allow_html=True)

    # Adding some space below the logos
    st.markdown("<br>", unsafe_allow_html=True)

    # Eligibility section (displayed only if Next button is clicked and all fields are filled)
    if st.session_state.get('show_eligibility', False):
        col1, col2 = st.columns((1, 5))
        options1 = ['Click here to select option', 'E36.00', 'F42.9']

        # Eligibility
        with col1:
            st.markdown(
                f'<div style="background-color: #3f3f3f; color: white; padding: 15px; border-radius: 100px; margin-bottom: 15px; width: 100%;" class="big-font-section">'
                '<strong>ELIGIBILITY</strong>'
                '<button class="info-button tooltip" id="info-btn">i'
                '<span class="tooltiptext">Confirm if the Economic Activity matches a suitable Macro Economic Sector as stipulated in the Delegated Acts or Technical Screening Criteria based on Nomenclature of Economic Activities(NACE) code</span>'
                '</div>', unsafe_allow_html=True)
        with col2:
            st.markdown(
                f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                '<strong>Provide the NACE (Nomenclature of Economic Activities) Code for Desalination</strong>'
                '<button class="info-button tooltip" id="info-btn">i'
                '<span class="tooltiptext">Desalination is not assigned a NACE code; however, desalination can be substituted by two other activity codes</span>'
                '</button>'
                '</div>', unsafe_allow_html=True)
            answer1 = st.selectbox('Select your response', options1, label_visibility='collapsed', key='answer1')

            answer2 = ""
            answer3 = ""
            answer4 = 0
            answer5 = 0
            answer6 = 0
            answer7 = 0
            answer8 = 0
            answer9 = []
            answer10 = ""
            answer11 = ""
            answer12 = ""
            if answer1 == 'E36.00':
                st.markdown(
                    f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                    '<strong>Is the desalination activity associated with construction, extension and operation of water collection, treatment and supply systems?</strong>'
                    '<button class="info-button tooltip" id="info-btn">i'
                    '<span class="tooltiptext">Desalination is not assigned a NACE code; however, desalination can be substituted by two other activity codes</span>'
                    '</button>'
                    '</div>', unsafe_allow_html=True)
                answer2 = st.radio('Yes or No?', options=['Yes', 'No'], label_visibility='collapsed')
            elif answer1 == 'Click here to select option':
                st.write("")
            else:
                st.markdown(
                    f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                    '<strong>NACE Code F42.9 is not eligible to continue.</strong>'
                    '</div>', unsafe_allow_html=True)

            if answer1 == "E36.00":
                st.markdown(
                    f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                    '<strong>Is the desalination activity associated with renewal of water collection, treatment and supply systems?</strong>'
                    '<button class="info-button tooltip" id="info-btn">i'
                    '<span class="tooltiptext">Desalination is not assigned a NACE code; however, desalination can be substituted by two other activity codes</span>'
                    '</button>'
                    '</div>', unsafe_allow_html=True)
                answer3 = st.radio('Yes or No?2', options=['Yes', 'No'], label_visibility='collapsed')
            if answer2 == "No" and answer3 == "No":
                st.markdown(
                    f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;text-align: center;" class="big-font">'
                    '<strong>The envisioned Activity is Not EU Taxonomy Eligible</strong>'
                    '</div>', unsafe_allow_html=True)
        col3, col4 = st.columns((1, 5))
        with col4:
            if (answer2 == "Yes" or answer3 == "Yes") and answer1 == "E36.00":
                st.markdown(
                    f'<div style="background-color: #00FF00; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                    '<strong>The envisioned Activity is EU Taxonomy Eligible</strong>'
                    '</div>', unsafe_allow_html=True)
        if answer2 == "Yes" or answer3 == "Yes":
            st.markdown("<hr class='bold-hr'>", unsafe_allow_html=True)
        col1, col2 = st.columns((1, 5))
        if answer2 == "Yes" or answer3 == "Yes":
            with col1:
                st.markdown(
                    f'<br><div style="background-color: #3f3f3f; color: white; padding: 15px; border-radius: 100px; margin-bottom: 15px; width: 100%;" class="big-font-section">'
                    '<strong>ALIGNMENT</strong>'
                    '<button class="info-button tooltip" id="info-btn">i'
                    '<span class="tooltiptext">Confirm if the Economic Activity provides a "Substantial Contribution" to one environmental objective and will "Do No Significant Harm" (DNSH)  to the remaining five environmental objectives, as designated per sector</span>'
                    '</div>', unsafe_allow_html=True)
            if answer2 == "Yes":
                # Alignment
                with col2:
                    st.markdown(
                        f'<br><div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                        '<strong>What is the net average energy consumption for abstraction and treatment for produced water supply (in kWh/m3)?</strong>'
                        '<button class="info-button tooltip" id="info-btn">i'
                        '<span class="tooltiptext">Should be less than or equal to 0.5</span>'
                        '</div>', unsafe_allow_html=True)
                    answer4 = st.number_input('Enter your response (kWh/m3)', min_value=0.0, step=0.01,label_visibility='collapsed')

                    answer5 = ""
                    st.markdown(
                        f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                        '<strong>What is the Leakage Level associated with this activity?</strong>'
                        '<button class="info-button tooltip" id="info-btn">i'
                        '<span class="tooltiptext">ILI= Current Annual Real Losses (CARL) / Unavoidable Annual Real Losses (UARL)<br>Should be less than or equal to 1.5</span>'
                        '</button>'
                        '</div>', unsafe_allow_html=True)
                    answer5 = st.number_input('Enter your response (kWh/m3)1', min_value=0.0, step=0.01,label_visibility='collapsed')

                    if answer5 > 1.5 or answer4 > 0.5:
                        st.markdown(
                            f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                            '<strong>The envisioned Activity is Not EU Taxonomy Eligible</strong>'
                            '</div>', unsafe_allow_html=True)
            if answer3 == "Yes" and answer5 <= 1.5 and answer4 <= 0.5:
                with col2:
                    st.markdown(
                        f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                        '<strong>What is the net average energy consumption compared to own baseline performance average for three years (in kWh/m3)</strong>'
                        '<button class="info-button tooltip" id="info-btn">i'
                        '<span class="tooltiptext">Should be greater than or equal to 20</span>'
                        '</button>'
                        '</div>', unsafe_allow_html=True)
                    answer6 = st.number_input('Enter your response (kWh/m3)', min_value=20.0, max_value=100.0, step=0.01,label_visibility='collapsed')

                    answer7 = ""
                    st.markdown(
                        f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                        '<strong>What is the Leakage Level between the current leakage level averaged over three years, calculated using the ILI of 1.5?</strong>'
                        '<button class="info-button tooltip" id="info-btn">i'
                        '<span class="tooltiptext">Should be greater than or equal to 20</span>'
                        '</button>'
                        '</div>', unsafe_allow_html=True)
                    answer7 = st.number_input('Enter your response (kWh/m3)1', min_value=20.0, max_value=100.0, step=0.01,label_visibility='collapsed')
            #2B
            if (answer2 == "Yes" and (answer5 < 1.5 and answer4 < 0.5) and (answer5 != 0.0 and answer4 != 0.0)) or (answer3 == "Yes"  and (answer6 >= 20 and answer7 >= 20)):
                with col2:
                    st.markdown(
                        f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                        '<strong>What is the Greenhouse Gas emissions of your activity (in CO2e/m3)?</strong>'
                        '<button class="info-button tooltip" id="info-btn">i'
                        '<span class="tooltiptext">Should be less than 1080</span>'
                        '</button>'
                        '</div>', unsafe_allow_html=True)
                    answer8 = st.number_input('Enter your response (gCO2e/m3)', min_value=0.0,max_value=1080.0, step=0.01,label_visibility='collapsed')

                    st.markdown(
                        f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                        '<strong>Does your activity significantly harm one of the remaining following Environmental Objectives:</strong>'
                        '</div>', unsafe_allow_html=True)
                    options = {
                            "Option 1": "Sustainable use and protection of water and marine resources",
                            "Option 2": "Transition to a circular economy",
                            "Option 3": "Pollution prevention and control",
                            "Option 4": "Protection and restoration of biodiversity and ecosystems",
                            "Option 5": "None of the above"
                        }
                    for key in options:
                        if key not in st.session_state:
                            st.session_state[key] = False

                    # Handle the logic for deselecting checkboxes before rendering the UI
                    if st.session_state["Option 5"]:
                        for k in options:
                            if k != "Option 5":
                                st.session_state[k] = False
                    else:
                        for k in options:
                            if k != "Option 5" and st.session_state[k]:
                                st.session_state["Option 5"] = False
                                break

                    # Create checkboxes
                    answer9 = {}
                    for key, label in options.items():
                        answer9[key] = st.checkbox(label, key=key)
                    answer9 = [label for key, label in options.items() if answer9[key]]
                if len(answer9) == 1 and ("None of the above" in answer9) and answer8 < 1080:
                    col3, col4 = st.columns((1, 5))
                    with col4:
                        st.markdown(
                        f'<div style="background-color: #00FF00; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                        '<strong>The envisioned Activity is EU Taxonomy Eligible</strong>'
                        '</div>', unsafe_allow_html=True)
                    st.markdown("<hr class='bold-hr'>", unsafe_allow_html=True)
                    col1, col2 = st.columns((1, 5))
                    with col1:
                        st.markdown(
                        f'<div style="background-color: #3f3f3f; color: white; padding: 15px; border-radius: 100px; margin-bottom: 15px; width: 100%;" class="big-font-section">'
                        '<strong>MINIMUM SAFEGUARDS</strong>'
                        '</div>', unsafe_allow_html=True)
                    #3A
                    with col2:
                        st.markdown(
                            f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                            '<strong>Did you implement an adequate Human Resources Due Diligence (HRDD)?</strong>'
                            '<button class="info-button tooltip" id="info-btn">i'
                            '<span class="tooltiptext">Promotion and protection of fundamental human rights as outlined in international conventions and agreements.</span>'
                            '</button>'
                            '</div>', unsafe_allow_html=True)
                        answer10 = st.radio('Yes or No?3', options=['Yes', 'No'], label_visibility='collapsed')
                        if answer10 == "No":
                            st.markdown(
                                f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                '<strong>Not Compliant</strong>'
                                '</div>', unsafe_allow_html=True)
                        else:
                            st.markdown(
                                f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                                '<strong>Is there any Signals of abuse to Human Resource?</strong>' 
                                '<button class="info-button tooltip" id="info-btn">i'
                                '<span class="tooltiptext">Support for fair labor practices, including ensuring safe working conditions, non- discrimination, and the right to organize and bargain collectively.</span>'
                                '</button>'
                                '</div>', unsafe_allow_html=True)
                            options = {
                                    "Option 1": "Found in breach of labour law or human rights",
                                    "Option 2": "Do not engage with relevant stakeholders",
                                    "Option 3": "None"
                                }
                            answer11 = {key: st.checkbox(label1) for key, label1 in options.items()}
                            answer11 = [label1 for key, label1 in options.items() if answer11[key]]
                            if len(answer11) >= 1 and "None" not in answer11:
                                st.markdown(
                                    f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                    '<strong>Not Compliant</strong>'
                                    '</div>', unsafe_allow_html=True)
                            elif len(answer11) == 1 and "None" in answer11:
                                #3B
                                st.markdown(
                                    f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                                    '<strong>Did any anti-corruption processes took place?</strong>'
                                    '<button class="info-button tooltip" id="info-btn">i'
                                    '<span class="tooltiptext">Implementation of measures to combat bribery and corruption in partner countries.</span>'
                                    '</button>'
                                    '</div>', unsafe_allow_html=True)
                                answer12 = st.radio('Yes or No?4', options=['Yes', 'No'], label_visibility='collapsed')
                                if answer12 == "Yes":
                                    st.markdown(
                                        f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                        '<strong>Not Compliant</strong>'
                                        '</div>', unsafe_allow_html=True)
                                else:
                                    st.markdown(
                                        f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                                        '<strong>Did the company or its senior management, including the senior management of its subsidiaries, has been convicted in court on corruption?</strong>'
                                        '<button class="info-button tooltip" id="info-btn">i'
                                        '<span class="tooltiptext">Assistance in establishing and enforcing anti-corruption policies and practices, aligned with international standards.</span>'
                                        '</button>'
                                        '</div>', unsafe_allow_html=True)
                                    answer13 = st.radio('Yes or No?9', options=['Yes', 'No'], label_visibility='collapsed')
                                    if answer13 == "Yes":
                                        st.markdown(
                                            f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                            '<strong>Not Compliant</strong>'
                                            '</div>', unsafe_allow_html=True)
                                    else:
                                        #3C
                                        st.markdown(
                                            f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                                            '<strong>The company confirms that it treats tax governance and compliance as important elements of oversight, and there are adequate tax risk management strategies and processes in place?</strong>'
                                            '<button class="info-button tooltip" id="info-btn">i'
                                            '<span class="tooltiptext">Collaboration on tax-related matters to promote fair and transparent tax systems.</span>'
                                            '</button>'
                                            '</div>', unsafe_allow_html=True)
                                        answer14 = st.radio('Yes or No?5', options=['Yes', 'No'], label_visibility='collapsed')
                                        if answer14 == "No":
                                            st.markdown(
                                                f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                                '<strong>Not Compliant</strong>'
                                                '</div>', unsafe_allow_html=True)
                                        else:
                                            st.markdown(
                                                f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                                                '<strong>Has the company or its subsidiaries violated the tax laws?</strong>'
                                                '<button class="info-button tooltip" id="info-btn">i'
                                                '<span class="tooltiptext">Support for partner countries in adopting international best practices in tax administration, compliance, and combating tax evasion.</span>'
                                                '</button>'
                                                '</div>', unsafe_allow_html=True)
                                            answer15 = st.radio('Yes or No?6', options=['Yes', 'No'], label_visibility='collapsed')
                                            if answer15 == "Yes":
                                                st.markdown(
                                                    f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                                    '<strong>Not Compliant</strong>'
                                                    '</div>', unsafe_allow_html=True)
                                            else:
                                                st.markdown(
                                                f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                                                '<strong>Does the company promote employee awareness on importance of compliance with all applicable competition laws and regulations?</strong>'
                                                '<button class="info-button tooltip" id="info-btn">i'
                                                '<span class="tooltiptext">Encouragement of open and competitive markets in partner countries to stimulate economic growth and development.</span>'
                                                '</button>'
                                                '</div>', unsafe_allow_html=True)
                                                answer16 = st.radio('Yes or No?7', options=['Yes', 'No'], label_visibility='collapsed')
                                                if answer16 == "No":
                                                    st.markdown(
                                                        f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                                        '<strong>Not Compliant</strong>'
                                                        '</div>', unsafe_allow_html=True)
                                                else:
                                                    st.markdown(
                                                    f'<div style="background-color: {bg_color}; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%;" class="big-font">'
                                                    '<strong>Does the company or its senior management, including the senior management of its subsidiaries, has been convicted on violating competition laws?</strong>'
                                                    '<button class="info-button tooltip" id="info-btn">i'
                                                    '<span class="tooltiptext">Assistance in establishing and enforcing competition laws and regulations to prevent anti-competitive behavior and monopolies .</span>'
                                                    '</button>'
                                                    '</div>', unsafe_allow_html=True)
                                                    answer17 = st.radio('Yes or No?8', options=['Yes', 'No'], label_visibility='collapsed')
                                                    if answer17 == "Yes":
                                                        st.markdown(
                                                            f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                                            '<strong>Not Compliant</strong>'
                                                            '</div>', unsafe_allow_html=True)
                                                    else:
                                                        st.markdown(
                                                            f'<div style="background-color: #00FF00; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                                                            '<strong>The envisioned Activity is compliant with the Minimum Safeguards of the EU Taxonomy. And the User can proceed to Financial Assessment.</strong>'
                                                            '</div>', unsafe_allow_html=True)
                                                        if st.session_state.page == 'main':
                                                            st.button("Continue", on_click=continue_to_phase)
                elif len(answer9) >= 1:
                    with col2:
                        st.markdown(
                            f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                            '<strong>EU Taxonomy does not align</strong>'
                            '</div>', unsafe_allow_html=True)
            else:
                with col2:
                    st.markdown(
                        f'<div style="background-color: #FF6347; color: white; padding: 15px; border-radius: 10px; margin-bottom: 15px; width: 100%; text-align: center;" class="big-font">'
                        '<strong>EU Taxonomy does not align</strong>'
                        '</div>', unsafe_allow_html=True)
    else:
        st.info("Please fill in all input fields to proceed with eligibility questions.")
#Make changes for the next page here
elif st.session_state.page == 'phase':

    df = load_financial_model(file_path,sheet_name='Inp_C',header=3)

    df_cleaned = df[df.notna().any(axis=1)]

    df_cleaned = df_cleaned.drop([1,3,5,8,13,18,23,29,31,35,38,41,46,48,54,56,62,70,72,77,83])
    df_cleaned = pd.DataFrame(df_cleaned)
    df_cleaned['Unit'] = df_cleaned['Unit'].fillna(0)
    df_cleaned['Unnamed: 4'] = df_cleaned['Unnamed: 4'].fillna(0)
    df_cleaned['Phase_1'] = df_cleaned['Phase_1'].fillna(0)
    df_cleaned['Phase_2'] = df_cleaned['Phase_2'].fillna(0)

    df_cleaned = df_cleaned.dropna(axis=1, how='any')

    # if 'df_cleaned' not in st.session_state:
    #     st.session_state.df_cleaned = load_financial_model(file_path,sheet_name='Inp_C',header=3)

    # df_cleaned = st.session_state.df_cleaned

    # st.write(df_cleaned,"after changes dataframe")

    def custom_number_input(label, key, placeholder, value=0.0):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", value)
        
        return st.number_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            step=0.01,
            on_change=handle_change,
            placeholder=placeholder
        )
    
    def custom_percentage_input(label, key, placeholder, value=0.0):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", value)
        
        return st.number_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            step=0.01,
            max_value=100.0,
            on_change=handle_change,
            placeholder=placeholder
        )

    def custom_text_input(label, key, placeholder):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", "")
        
        return st.text_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            on_change=handle_change,
            placeholder=placeholder
        )
    
    def format_date(date):
        """Format date to dd/mm/yy."""
        return date.strftime('%d/%m/%Y')
    
    def custom_date_input(label, key):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", datetime.now().date())
        
        if key not in st.session_state:
            st.session_state[key] = datetime.now().date()
        
        return st.date_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            on_change=handle_change
        )
    

    def calculate_future_date_months(date, months):
        whole_months = int(months)
        fractional_months = months - whole_months
        future_date = date + relativedelta(months=whole_months)
        additional_days = int(fractional_months * 30)
        future_date += relativedelta(days=additional_days)
        return future_date
    
    def calculate_future_date_years(date, years):
        whole_years = int(years)
        fractional_years = years - whole_years
        future_date = date + relativedelta(years=whole_years)
        additional_days = int(fractional_years * 365)
        future_date += relativedelta(days=additional_days)
        return future_date

    st.markdown(
        f'<div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 15px; margin-bottom: 15px; width: 100%; text-align: center; font-size: 54px; margin-top: -50px;" class="center-text">'
        '<strong>FINANCIAL ASSESSMENT</strong>'
        '</div>', unsafe_allow_html=True)
    
    # Only show phase selection if it hasn't been selected yet
    if st.session_state.selected_phase_option == "Select an option":
        options = ["Select an option", "1", "2"]
        selected = st.selectbox(
            "Is your project a one phase or two phase project?", 
            options, 
            key='100'
        )
        # Update the session state immediately after selection
        if selected != "Select an option":
            st.session_state.selected_phase_option = selected
            st.rerun()  # Rerun to update the UI immediately
    
    if st.session_state.selected_phase_option == "1":
        check_phase1 = True
        check_phase2 = False
    elif st.session_state.selected_phase_option == "2":
        check_phase1 = True
        check_phase2 = True
    else:
        check_phase1 = False
        check_phase2 = False

    if check_phase1:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            f'<div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 15px; margin-bottom: 15px; width: 80%; text-align: center; font-size: 36px; margin-top: -50px;" class="center-text">'
            '<strong>Phase 1</strong>'
            '</div>', unsafe_allow_html=True)
        
        # Use unique keys for each input to avoid conflicts
        fcp1 = custom_date_input("Financial Close - Phase 1", 'fcp1')
        field7 = custom_date_input("Construction Start Date - Phase 1", 'field7')
        field8 = custom_number_input("Construction Period(in months) - Phase 1", 'field8', "Enter", st.session_state.field8)
        
        if field8 != 0:    
            construction_end_date = calculate_future_date_months(field7, field8)
            st.write("Construction End Date: - Phase 1", construction_end_date)
            st.session_state.field9 = construction_end_date
            
            operations_start_date = construction_end_date + timedelta(days=1)
            st.session_state.osdp1 = operations_start_date
            st.write("Operations Start Date - Phase 1:", operations_start_date)
        
        field10 = custom_number_input("Operations Period(in Years) - Phase 1", 'field10', "Enter", st.session_state.field10)
        
        if field10 != 0 and st.session_state.osdp1:
            operations_end_date = calculate_future_date_years(st.session_state.osdp1, field10)
            st.write("Operations End Date: - Phase 1", operations_end_date)
            st.session_state.oedp1 = operations_end_date
        
        if st.session_state.osdp1:
            st.session_state.drsdp1 = st.session_state.osdp1
            st.write("Debt Repayment Start Date - Phase 1:", st.session_state.drsdp1)
        
        drtp1 = custom_number_input("Debt Repayment Tenor(in Years) - Phase 1", 'drtp1', "Enter", st.session_state.drtp1)
        
        if drtp1 != 0 and st.session_state.drsdp1:
            debt_repayment_end_date = calculate_future_date_years(st.session_state.drsdp1, drtp1)
            st.write("Debt Repayment End Date: - Phase 1", debt_repayment_end_date)
            st.session_state.dredp1 = debt_repayment_end_date
        
        cepsp1 = custom_number_input("Capital Expenditure - Pre sensitivity(in LE'000 or $'000) - Phase 1", 'cepsp1', "Enter", st.session_state.cepsp1)
        ces1 = custom_percentage_input("Capital Expenditure - Sensitivity (%) - Phase 1", 'ces1', "Enter", st.session_state.ces1)
        
        if cepsp1 != 0 and ces1 != 0:
            post_sensitivity_value = cepsp1 * (ces1 / 100)
            st.write("Calculated Capital Expenditure - Post Sensitivity (in LE'000 or $'000):", post_sensitivity_value)
            st.session_state.field11 = post_sensitivity_value
        
        field12 = custom_percentage_input("Debt (%) - Phase 1", 'field12', "Enter", st.session_state.field12)
        
        if field12 != 0:
            equity_percentage = 1 - (field12 / 100)
            round_equity = round(equity_percentage * 100)
            st.write("Equity (%): ", round_equity)
            st.session_state.cep1 = round_equity
        
        upfp1 = custom_number_input("Upfront Fees (%) - Phase 1", 'upfp1', "Enter", st.session_state.upfp1)
        cfp1 = custom_number_input("Commitment Fees (%) - Phase 1", 'cfp1', "Enter", st.session_state.cfp1)
        field14 = custom_percentage_input("Construction Interest Rate (Base Rate %) - Phase 1", 'field14', "Enter", st.session_state.field14)
        field15 = custom_percentage_input("Construction Interest Rate (Margin Spread %) - Phase 1", 'field15', "Enter", st.session_state.field15)
        
        if field14 != 0 and field15 != 0:
            all_in_rate = field14 + field15
            st.write("All in Rate (%)", all_in_rate)
            st.session_state.field16 = all_in_rate
        
        field17 = custom_percentage_input("Operations Interest Rate (Base Rate %) - Phase 1", 'field17', "Enter", st.session_state.field17)
        field18 = custom_percentage_input("Operations Interest Rate (Margin Spread %) - Phase 1", 'field18', "Enter", st.session_state.field18)
        
        if field17 != 0 and field18 != 0:
            all_in_rate_ops = field17 + field18
            st.write("All in Rate (%) - Phase 1", all_in_rate_ops)
            st.session_state.field19 = all_in_rate_ops
        
        ofwaccp1 = custom_percentage_input("Offtake - WACC (%) - Phase 1", "ofwaccp1", "Enter", st.session_state.ofwaccp1)
        drcitrp1 = custom_percentage_input("Corporate Income Tax Rate (%) - Phase 1", "drcitrp1", "Enter", st.session_state.drcitrp1)

        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            f'<div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 15px; margin-bottom: 15px; width: 80%; text-align: center; font-size: 36px; margin-top: -50px;" class="center-text">'
            '<strong>Phase 1</strong>'
            '</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns((1, 1))
        with col1:
            st.markdown(
                '<div style="font-size:24px; font-weight:bold; color:#333333; border-bottom:2px solid #cccccc; padding-bottom:8px; margin-bottom:15px;">'
                'Fixed Operating Costs Per Year - Phase 1'
                '</div>',
                unsafe_allow_html=True
            )
            cgp1 = custom_number_input("Chlorine Gas (in LE'000 or $'000) - Phase 1", 'cgp1', "Enter", st.session_state.cgp1)
            cfltp1 = custom_number_input("Chemical for laboratory test (in LE'000 or $'000) - Phase 1", 'cfltp1', "Enter", st.session_state.cfltp1)
            ogsgp1 = custom_number_input("Fuel (Oil, Gas, Solar and Gasoline (in LE'000 or $'000) - Phase 1", 'ogsgp1', "Enter", st.session_state.ogsgp1)
            ofo1 = custom_number_input("Other Fixed Opex (in LE'000 or $'000) - Phase 1", 'ofo1', "Enter", st.session_state.ofo1)
        
        with col2:
            st.markdown(
                '<div style="font-size:24px; font-weight:bold; color:#333333; border-bottom:2px solid #cccccc; padding-bottom:8px; margin-bottom:15px;">'
                'Variable Operating Costs Per Year - Phase 1'
                '</div>',
                unsafe_allow_html=True
            )
            field25 = custom_number_input("Labor (in LE'000 or $'000) - Phase 1", 'field25', "Enter", st.session_state.field25)
            field26 = custom_number_input("Spare Part Cost (in LE'000 or $'000) - Phase 1", 'field26', "Enter", st.session_state.field26)
            field27 = custom_number_input("Energy Costs (LE or $/Kw) - Phase 1", 'field27', "Enter", st.session_state.field27)
            field28 = custom_number_input("Energy Consumption (KW/m³) - Phase 1", 'field28', "Enter", st.session_state.field28)
            
            if field27 != 0 and field28 != 0:
                effective_price = field27 * field28
                st.write("Effective Price- Energy Costs (LE/m³) - Phase 1", effective_price)
                st.session_state.field29 = effective_price
            
            field34 = custom_number_input("Maintenance Costs (in LE'000 or $'000) - Phase 1", 'field34', "Enter", st.session_state.field34)

        if check_phase1 and not check_phase2:
            st.button("Continue to Risk Management", on_click=continue_to_risk_management,key='cont1')
        if check_phase2:
            st.button("Continue", on_click=continue_to_phase2,key='cont2')


        # Updating data into financial model:

        # if 'oedp1' not in st.session_state:
        #      st.session_state.oedp1 = None

        # if 'dredp1' not in st.session_state:
        #      st.session_state.dredp1 = None

        # if 'cep1' not in st.session_state:
        #      st.session_state.cep1 = None

        # if 'field29' not in st.session_state:
        #      st.session_state.field29 = None

        # if 'osdp1' not in st.session_state:
        #      st.session_state.osdp1 = None

        # if 'drsdp1' not in st.session_state:
        #      st.session_state.drsdp1 = None

        if fcp1:
            # Store the new date input into the DataFrame
            df_cleaned.at[6, 'Phase_1'] = fcp1

        if field7:
            df_cleaned.at[9, 'Phase_1'] = field7

        if field8:
            df_cleaned.at[10, 'Phase_1'] = field8

        # if field9:
        #     df_cleaned.at[11, 'Phase_1'] = field9

        # if getattr(st.session_state, 'osdp1', None):
        #     df_cleaned.at[14, 'Phase_1'] = st.session_state.osdp1

        if field10:
            df_cleaned.at[15, 'Phase_1'] = field10

        # if getattr(st.session_state, 'oedp1', None):
        #     df_cleaned.at[16, 'Phase_1'] = st.session_state.oedp1

        # if getattr(st.session_state, 'drsdp1', None):
        #     df_cleaned.at[19, 'Phase_1'] = st.session_state.drsdp1

        if drtp1:
            df_cleaned.at[20, 'Phase_1'] = drtp1

        # if getattr(st.session_state, 'dredp1', None):
        #     df_cleaned.at[21, 'Phase_1'] = st.session_state.dredp1

        if cepsp1:
            df_cleaned.at[25, 'Phase_1'] = cepsp1

        if ces1:
            df_cleaned.at[26, 'Phase_1'] = ces1
            
        # if st.session_state.field11:
        #     df_cleaned.at[27, 'Phase_1'] = st.session_state.field11

        if field12:
            df_cleaned.at[32, 'Phase_1'] = field12

        # if getattr(st.session_state, 'cep1', None):
        #     df_cleaned.at[33, 'Phase_1'] = st.session_state.cep1

        if upfp1:
            df_cleaned.at[36, 'Phase_1'] = upfp1

        if cfp1:
            df_cleaned.at[39, 'Phase_1'] = cfp1

        if field14:
            df_cleaned.at[42, 'Phase_1'] = field14

        if field15:
            df_cleaned.at[43, 'Phase_1'] = field15

        if cgp1:
            df_cleaned.at[57, 'Phase_1'] = cgp1

        if cfltp1:
            df_cleaned.at[58, 'Phase_1'] = cfltp1

        if ogsgp1:
            df_cleaned.at[59, 'Phase_1'] = ogsgp1

        if ofo1:
            df_cleaned.at[60, 'Phase_1'] = ofo1

        if field25:
            df_cleaned.at[63, 'Phase_1'] = field25

        if field26:
            df_cleaned.at[64, 'Phase_1'] = field26

        if field27:
            df_cleaned.at[65, 'Phase_1'] = field27

        if field28:
            df_cleaned.at[66, 'Phase_1'] = field28

        # if getattr(st.session_state, 'field29', None):
        #     df_cleaned.at[67, 'Phase_1'] = st.session_state.field29

        if field34:
            df_cleaned.at[68, 'Phase_1'] = field34

        if field17:
            df_cleaned.at[73, 'Phase_1'] = field17

        if field18:
            df_cleaned.at[74, 'Phase_1'] = field18

        # if st.session_state.field19:
        #     df_cleaned.at[75, 'Phase_1'] = st.session_state.field19

        if ofwaccp1:
            df_cleaned.at[79, 'Phase_1'] = ofwaccp1

        if drcitrp1:
            df_cleaned.at[81, 'Phase_1'] = drcitrp1


        # st.write(df_cleaned,'----------------updating')


        file_extension = Path(file_path).suffix.lower()[1:]
        if file_extension in ['xlsx', 'xls']:
            with open(file_path, 'rb') as file:
                content = file.read(4)
                workbook = load_workbook(file_path)
                sheet = workbook['Inp_C']


        if 'Phase_1' in df_cleaned.columns:
            sheet.cell(row=11, column=10, value=df_cleaned.at[6, 'Phase_1'])
            sheet.cell(row=14, column=10, value=df_cleaned.at[9, 'Phase_1'])
            sheet.cell(row=15, column=10, value=df_cleaned.at[10, 'Phase_1']) 
            # sheet.cell(row=16, column=10, value=df_cleaned.at[11, 'Phase_1']) 
            # sheet.cell(row=19, column=10, value=df_cleaned.at[14, 'Phase_1']) 
            sheet.cell(row=20, column=10, value=df_cleaned.at[15, 'Phase_1']) 
            # sheet.cell(row=21, column=10, value=df_cleaned.at[16, 'Phase_1']) 
            # sheet.cell(row=24, column=10, value=df_cleaned.at[19, 'Phase_1'])  
            sheet.cell(row=25, column=10, value=df_cleaned.at[20, 'Phase_1']) 
            # sheet.cell(row=26, column=10, value=df_cleaned.at[21, 'Phase_1']) 
            sheet.cell(row=30, column=10, value=df_cleaned.at[25, 'Phase_1']) 
            sheet.cell(row=31, column=10, value=df_cleaned.at[26, 'Phase_1']  / 100) 
            # sheet.cell(row=32, column=10, value=df_cleaned.at[27, 'Phase_1'])  
            sheet.cell(row=37, column=10, value=df_cleaned.at[32, 'Phase_1'] / 100) 
            # sheet.cell(row=38, column=10, value=df_cleaned.at[33, 'Phase_1'] / 100) 
            sheet.cell(row=41, column=10, value=df_cleaned.at[36, 'Phase_1'] / 100) 
            sheet.cell(row=44, column=10, value=df_cleaned.at[39, 'Phase_1'] / 100) 
            sheet.cell(row=47, column=10, value=df_cleaned.at[42, 'Phase_1'] / 100) 
            sheet.cell(row=48, column=10, value=df_cleaned.at[43, 'Phase_1'] / 100) 
            sheet.cell(row=62, column=10, value=df_cleaned.at[57, 'Phase_1']) 
            sheet.cell(row=63, column=10, value=df_cleaned.at[58, 'Phase_1']) 
            sheet.cell(row=64, column=10, value=df_cleaned.at[59, 'Phase_1']) 
            sheet.cell(row=65, column=10, value=df_cleaned.at[60, 'Phase_1']) 
            sheet.cell(row=68, column=10, value=df_cleaned.at[63, 'Phase_1']) 
            sheet.cell(row=69, column=10, value=df_cleaned.at[64, 'Phase_1']) 
            sheet.cell(row=70, column=10, value=df_cleaned.at[65, 'Phase_1']) 
            sheet.cell(row=71, column=10, value=df_cleaned.at[66, 'Phase_1']) 
            # sheet.cell(row=72, column=10, value=df_cleaned.at[67, 'Phase_1']) 
            sheet.cell(row=73, column=10, value=df_cleaned.at[68, 'Phase_1']) 
            sheet.cell(row=78, column=10, value=df_cleaned.at[73, 'Phase_1'] / 100)
            sheet.cell(row=79, column=10, value=df_cleaned.at[74, 'Phase_1'] / 100)
            # sheet.cell(row=80, column=10, value=df_cleaned.at[75, 'Phase_1'] / 100)
            sheet.cell(row=84, column=10, value=df_cleaned.at[79, 'Phase_1']  / 100)
            sheet.cell(row=86, column=10, value=df_cleaned.at[81, 'Phase_1']  / 100)

            # workbook.save(file_path)


            st.session_state.df_cleaned_phase_1 = df_cleaned

            # st.write(df_cleaned,"updated data frame")

            # st.write(st.session_state.df_cleaned_phase_1)

 
elif st.session_state.page == 'phase2':

    # df = load_financial_model(file_path,sheet_name='Inp_C',header=3)


    # df_cleaned = df[df.notna().any(axis=1)]

    # # more_df_cleaned = df_cleaned.drop([1,3,5,8,13,18,23,29,31,35,38,41,46,48,54,56,62,70,72,77,83,85,86,87])
    # df_cleaned = df_cleaned.drop([1,3,5,8,13,18,23,29,31,35,38,41,46,48,54,56,62,70,72,77,83])

    # df_cleaned['Unit'] = df_cleaned['Unit'].fillna(0)
    # df_cleaned['Unnamed: 4'] = df_cleaned['Unnamed: 4'].fillna(0)
    # df_cleaned['Phase_1'] = df_cleaned['Phase_1'].fillna(0)
    # df_cleaned['Phase_2'] = df_cleaned['Phase_2'].fillna(0)

    # df_cleaned = df_cleaned.dropna(axis=1, how='any')


    if 'df_cleaned_phase_1' in st.session_state:
        df_cleaned_phase_2 = st.session_state.df_cleaned_phase_1
    else:
        st.write("Phase 1 data is not available. Please update Phase 1 first.")

    def custom_number_input(label, key, placeholder, value=0.0):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", value)
        
        return st.number_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            step=0.01,
            on_change=handle_change,
            placeholder=placeholder
        )
    
    def custom_percentage_input(label, key, placeholder, value=0.0):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", value)
        
        return st.number_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            step=0.01,
            max_value=100.0,
            on_change=handle_change,
            placeholder=placeholder
        )

    def custom_text_input(label, key, placeholder):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", "")
        
        return st.text_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            on_change=handle_change,
            placeholder=placeholder
        )
    
    def custom_date_input(label, key):
        def handle_change():
            st.session_state[key] = st.session_state.get(f"{key}_input", datetime.now().date())
        
        if key not in st.session_state:
            st.session_state[key] = datetime.now().date()
        
        return st.date_input(
            label,
            key=f"{key}_input",
            value=st.session_state[key],
            on_change=handle_change
        )
    
    def calculate_future_date_months(date, months):
        whole_months = int(months)
        fractional_months = months - whole_months
        future_date = date + relativedelta(months=whole_months)
        additional_days = int(fractional_months * 30)
        future_date += relativedelta(days=additional_days)
        return future_date
    
    def calculate_future_date_years(date, years):
        whole_years = int(years)
        fractional_years = years - whole_years
        future_date = date + relativedelta(years=whole_years)
        additional_days = int(fractional_years * 365)
        future_date += relativedelta(days=additional_days)
        return future_date

    st.markdown(
        f'<div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 15px; margin-bottom: 15px; width: 100%; text-align: center; font-size: 54px; margin-top: -50px;" class="center-text">'
        '<strong>FINANCIAL ASSESSMENT</strong>'
        '</div>', unsafe_allow_html=True)
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown(
        f'<div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 15px; margin-bottom: 15px; width: 80%; text-align: center; font-size: 36px; margin-top: -50px;" class="center-text">'
        '<strong>Phase 2</strong>'
        '</div>', unsafe_allow_html=True)
    
    fcp2 = custom_date_input("Financial Close - Phase 2", 'fcp2')
    field7p2 = custom_date_input("Construction Start Date - Phase 2", 'field7p2')
    field8p2 = custom_number_input("Construction Period(in months) - Phase 2", 'field8p2', "Enter", st.session_state.field8p2)
    
    if field8p2 != 0:    
        construction_end_date = calculate_future_date_months(field7p2, field8p2)
        st.write("Construction End Date: - Phase 2", construction_end_date)
        st.session_state.field9p2 = construction_end_date
        
        operations_start_date = construction_end_date + timedelta(days=1)
        st.session_state.osdp2 = operations_start_date
        st.write("Operations Start Date - Phase 2:", operations_start_date)
    
    field10p2 = custom_number_input("Operations Period(in Years) - Phase 2", 'field10p2', "Enter", st.session_state.field10p2)
    
    if field10p2 != 0 and st.session_state.osdp2:
        operations_end_date = calculate_future_date_years(st.session_state.osdp2, field10p2)
        st.write("Operations End Date: - Phase 2", operations_end_date)
        st.session_state.oedp2 = operations_end_date
    
    if st.session_state.osdp2:
        st.session_state.drsdp2 = st.session_state.osdp2
        st.write("Debt Repayment Start Date - Phase 2:", st.session_state.drsdp2)
    
    drtp2 = custom_number_input("Debt Repayment Tenor(in Years) - Phase 2", 'drtp2', "Enter", st.session_state.drtp2)
    
    if drtp2 != 0 and st.session_state.drsdp2:
        debt_repayment_end_date = calculate_future_date_years(st.session_state.drsdp2, drtp2)
        st.write("Debt Repayment End Date: - Phase 2", debt_repayment_end_date)
        st.session_state.dredp2 = debt_repayment_end_date
    
    cepsp2 = custom_number_input("Capital Expenditure - Pre sensitivity(in LE'000 or $'000) - Phase 2", 'cepsp2', "Enter", st.session_state.cepsp2)
    ces2 = custom_percentage_input("Capital Expenditure - Sensitivity (%) - Phase 2", 'ces2', "Enter", st.session_state.ces2)
    
    if cepsp2 != 0 and ces2 != 0:
        post_sensitivity_value = cepsp2 * (ces2 / 100)
        st.write("Calculated Capital Expenditure - Post Sensitivity (in LE'000 or $'000):", post_sensitivity_value)
        st.session_state.field11p2 = post_sensitivity_value
    
    field12p2 = custom_percentage_input("Debt (%) - Phase 2", 'field12p2', "Enter", st.session_state.field12p2)
    
    if field12p2 != 0:
        equity_percentage = 1 - (field12p2 / 100)
        round_equity = round(equity_percentage * 100)
        st.write("Equity (%): ", round_equity)
        st.session_state.cep2 = round_equity
    
    upfp2 = custom_number_input("Upfront Fees (%) - Phase 2", 'upfp2', "Enter", st.session_state.upfp2)
    cfp2 = custom_number_input("Commitment Fees (%) - Phase 2", 'cfp2', "Enter", st.session_state.cfp2)
    field14p2 = custom_percentage_input("Construction Interest Rate (Base Rate %) - Phase 2", 'field14p2', "Enter", st.session_state.field14p2)
    field15p2 = custom_percentage_input("Construction Interest Rate (Margin Spread %) - Phase 2", 'field15p2', "Enter", st.session_state.field15p2)
    
    if field14p2 != 0 and field15p2 != 0:
        all_in_rate = field14p2 + field15p2
        st.write("All in Rate (%)", all_in_rate)
        st.session_state.field16p2 = all_in_rate
    
    field17p2 = custom_percentage_input("Operations Interest Rate (Base Rate %) - Phase 2", 'field17p2', "Enter", st.session_state.field17p2)
    field18p2 = custom_percentage_input("Operations Interest Rate (Margin Spread %) - Phase 2", 'field18p2', "Enter", st.session_state.field18p2)
    
    if field17p2 != 0 and field18p2 != 0:
        all_in_rate_ops = field17p2 + field18p2
        st.write("All in Rate (%) - Phase 2", all_in_rate_ops)
        st.session_state.field19p2 = all_in_rate_ops
    
    ofwaccp2 = custom_percentage_input("Offtake - WACC (%) - Phase 2", "ofwaccp2", "Enter", st.session_state.ofwaccp2)
    drcitrp2 = custom_percentage_input("Corporate Income Tax Rate (%) - Phase 2", "drcitrp2", "Enter", st.session_state.drcitrp2)

    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown(
        f'<div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 15px; margin-bottom: 15px; width: 80%; text-align: center; font-size: 36px; margin-top: -50px;" class="center-text">'
        '<strong>Phase 2</strong>'
        '</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns((1, 1))
    with col1:
        st.markdown(
            '<div style="font-size:24px; font-weight:bold; color:#333333; border-bottom:2px solid #cccccc; padding-bottom:8px; margin-bottom:15px;">'
            'Fixed Operating Costs Per Year - Phase 2'
            '</div>',
            unsafe_allow_html=True
        )
        cgp2 = custom_number_input("Chlorine Gas (in LE'000 or $'000) - Phase 2", 'cgp2', "Enter", st.session_state.cgp2)
        cfltp2 = custom_number_input("Chemical for laboratory test (in LE'000 or $'000) - Phase 2", 'cfltp2', "Enter", st.session_state.cfltp2)
        ogsgp2 = custom_number_input("Fuel (Oil, Gas, Solar and Gasoline (in LE'000 or $'000) - Phase 2", 'ogsgp2', "Enter", st.session_state.ogsgp2)
        ofo2 = custom_number_input("Other Fixed Opex (in LE'000 or $'000) - Phase 2", 'ofo2', "Enter", st.session_state.ofo2)
    
    with col2:
        st.markdown(
            '<div style="font-size:24px; font-weight:bold; color:#333333; border-bottom:2px solid #cccccc; padding-bottom:8px; margin-bottom:15px;">'
            'Variable Operating Costs Per Year - Phase 2'
            '</div>',
            unsafe_allow_html=True
        )
        field25p2 = custom_number_input("Labor (in LE'000 or $'000) - Phase 2", 'field25p2', "Enter", st.session_state.field25p2)
        field26p2 = custom_number_input("Spare Part Cost (in LE'000 or $'000) - Phase 2", 'field26p2', "Enter", st.session_state.field26p2)
        field27p2 = custom_number_input("Energy Costs (LE or $/Kw) - Phase 2", 'field27p2', "Enter", st.session_state.field27p2)
        field28p2 = custom_number_input("Energy Consumption (KW/m³) - Phase 2", 'field28p2', "Enter", st.session_state.field28p2)
        
        if field27p2 != 0 and field28p2 != 0:
            effective_price = field27p2 * field28p2
            st.write("Effective Price- Energy Costs (LE/m³) - Phase 2", effective_price)
            st.session_state.field29p2 = effective_price
        
        field34p2 = custom_number_input("Maintenance Costs (in LE'000 or $'000) - Phase 2", 'field34p2', "Enter", st.session_state.field34p2)

    # Add navigation buttons
    col1, col2 = st.columns([1, 18])
    with col1:
        st.button("Back", on_click=go_back_to_phase)
    with col2:
        st.button("Continue to Risk Assessment", on_click=continue_to_risk_management)

    # Update DataFrame
    if fcp2:
        df_cleaned_phase_2.at[6, 'Phase_2'] = fcp2

    if field7p2:
        df_cleaned_phase_2.at[9, 'Phase_2'] = field7p2

    if field8p2:
        df_cleaned_phase_2.at[10, 'Phase_2'] = field8p2

    # if field9p2:
    #     df_cleaned_phase_2.at[11, 'Phase_2'] = field9p2

    # if getattr(st.session_state, 'osdp2', None):
    #     df_cleaned_phase_2.at[14, 'Phase_2'] = st.session_state.osdp2

    if field10p2:
        df_cleaned_phase_2.at[15, 'Phase_2'] = field10p2

    # if getattr(st.session_state, 'oedp2', None):
    #     df_cleaned_phase_2.at[16, 'Phase_2'] = st.session_state.oedp2

    # if getattr(st.session_state, 'drsdp2', None):
    #     df_cleaned_phase_2.at[19, 'Phase_2'] = st.session_state.drsdp2

    if drtp2:
        df_cleaned_phase_2.at[20, 'Phase_2'] = drtp2

    # if getattr(st.session_state, 'dredp2', None):
    #     df_cleaned_phase_2.at[21, 'Phase_2'] = st.session_state.dredp2

    if cepsp2:
        df_cleaned_phase_2.at[25, 'Phase_2'] = cepsp2

    if ces2:
        df_cleaned_phase_2.at[26, 'Phase_2'] = ces2
        
    # if getattr(st.session_state, 'field69', None):
    #     df_cleaned_phase_2.at[27, 'Phase_2'] = st.session_state.field69

    if st.session_state.field38:
        df_cleaned_phase_2.at[32, 'Phase_2'] = st.session_state.field38

    # if getattr(st.session_state, 'cep2', None):
    #     df_cleaned_phase_2.at[33, 'Phase_2'] = st.session_state.cep2

    if st.session_state.upfp2:
        df_cleaned_phase_2.at[36, 'Phase_2'] = st.session_state.upfp2

    if st.session_state.cfp2:
        df_cleaned_phase_2.at[39, 'Phase_2'] = st.session_state.cfp2

    if st.session_state.field42:
        df_cleaned_phase_2.at[42, 'Phase_2'] = st.session_state.field42

    if st.session_state.field43:
        df_cleaned_phase_2.at[43, 'Phase_2'] = st.session_state.field43

    # if st.session_state.field44:
    #     df_cleaned_phase_2.at[44, 'Phase_2'] = st.session_state.field44

    if st.session_state.cgp2:
        df_cleaned_phase_2.at[57, 'Phase_2'] = st.session_state.cgp2

    if st.session_state.cfltp2:
        df_cleaned_phase_2.at[58, 'Phase_2'] = st.session_state.cfltp2

    if st.session_state.ogsgp2:
        df_cleaned_phase_2.at[59, 'Phase_2'] = st.session_state.ogsgp2

    if st.session_state.ofo2:
        df_cleaned_phase_2.at[60, 'Phase_2'] = st.session_state.ofo2

    if st.session_state.field53:
        df_cleaned_phase_2.at[63, 'Phase_2'] = st.session_state.field53

    if st.session_state.field62:
        df_cleaned_phase_2.at[64, 'Phase_2'] = st.session_state.field62

    if st.session_state.field58:
        df_cleaned_phase_2.at[65, 'Phase_2'] = st.session_state.field58

    if st.session_state.field59:
        df_cleaned_phase_2.at[66, 'Phase_2'] = st.session_state.field59

    # if st.session_state.field60:
    #     df_cleaned_phase_2.at[67, 'Phase_2'] = st.session_state.field60

    if st.session_state.field54:
        df_cleaned_phase_2.at[68, 'Phase_2'] = st.session_state.field54

    if st.session_state.field45:
        df_cleaned_phase_2.at[73, 'Phase_2'] = st.session_state.field45

    if st.session_state.field46:
        df_cleaned_phase_2.at[74, 'Phase_2'] = st.session_state.field46

    # if st.session_state.field47:
    #     df_cleaned_phase_2.at[75, 'Phase_2'] = st.session_state.field47

    if st.session_state.ofwaccp2:
        df_cleaned_phase_2.at[79, 'Phase_2'] = st.session_state.ofwaccp2

    if st.session_state.drcitrp2:
        df_cleaned_phase_2.at[81, 'Phase_2'] = st.session_state.drcitrp2



    file_extension = Path(file_path).suffix.lower()[1:]
    if file_extension in ['xlsx', 'xls']:
        workbook = load_workbook(file_path)
        sheet = workbook['Inp_C']

    if 'Phase_2' in df_cleaned_phase_2.columns:
        sheet.cell(row=11, column=11, value=df_cleaned_phase_2.at[6, 'Phase_2'])
        sheet.cell(row=14, column=11, value=df_cleaned_phase_2.at[9, 'Phase_2'])
        sheet.cell(row=15, column=11, value=df_cleaned_phase_2.at[10, 'Phase_2']) 
        # sheet.cell(row=16, column=11, value=df_cleaned_phase_2.at[11, 'Phase_2']) 
        # sheet.cell(row=19, column=11, value=df_cleaned_phase_2.at[14, 'Phase_2']) 
        sheet.cell(row=20, column=11, value=df_cleaned_phase_2.at[15, 'Phase_2']) 
        # sheet.cell(row=21, column=11, value=df_cleaned_phase_2.at[16, 'Phase_2']) 
        # sheet.cell(row=24, column=11, value=df_cleaned_phase_2.at[19, 'Phase_2']) 
        sheet.cell(row=25, column=11, value=df_cleaned_phase_2.at[20, 'Phase_2']) 
        # sheet.cell(row=26, column=11, value=df_cleaned_phase_2.at[21, 'Phase_2']) 
        sheet.cell(row=30, column=11, value=df_cleaned_phase_2.at[25, 'Phase_2']) 
        sheet.cell(row=31, column=11, value=df_cleaned_phase_2.at[26, 'Phase_2']  / 100) 
        # sheet.cell(row=32, column=11, value=df_cleaned_phase_2.at[27, 'Phase_2']) 
        sheet.cell(row=37, column=11, value=df_cleaned_phase_2.at[32, 'Phase_2'] / 100) 
        # sheet.cell(row=38, column=11, value=df_cleaned_phase_2.at[33, 'Phase_2'] / 100) 
        sheet.cell(row=41, column=11, value=df_cleaned_phase_2.at[36, 'Phase_2'] / 100) 
        sheet.cell(row=44, column=11, value=df_cleaned_phase_2.at[39, 'Phase_2'] / 100) 
        sheet.cell(row=47, column=11, value=df_cleaned_phase_2.at[42, 'Phase_2'] / 100) 
        sheet.cell(row=48, column=11, value=df_cleaned_phase_2.at[43, 'Phase_2'] / 100) 
        sheet.cell(row=62, column=11, value=df_cleaned_phase_2.at[57, 'Phase_2']) 
        sheet.cell(row=63, column=11, value=df_cleaned_phase_2.at[58, 'Phase_2'])   
        sheet.cell(row=64, column=11, value=df_cleaned_phase_2.at[59, 'Phase_2']) 
        sheet.cell(row=65, column=11, value=df_cleaned_phase_2.at[60, 'Phase_2']) 
        sheet.cell(row=68, column=11, value=df_cleaned_phase_2.at[63, 'Phase_2']) 
        sheet.cell(row=69, column=11, value=df_cleaned_phase_2.at[64, 'Phase_2']) 
        sheet.cell(row=70, column=11, value=df_cleaned_phase_2.at[65, 'Phase_2']) 
        sheet.cell(row=71, column=11, value=df_cleaned_phase_2.at[66, 'Phase_2']) 
        # sheet.cell(row=72, column=11, value=df_cleaned_phase_2.at[67, 'Phase_2']) 
        sheet.cell(row=78, column=11, value=df_cleaned_phase_2.at[73, 'Phase_2'] / 100)
        sheet.cell(row=79, column=11, value=df_cleaned_phase_2.at[74, 'Phase_2'] / 100)
        # sheet.cell(row=80, column=11, value=df_cleaned_phase_2.at[75, 'Phase_2'] / 100)
        sheet.cell(row=84, column=11, value=df_cleaned_phase_2.at[79, 'Phase_2']  / 100)
        sheet.cell(row=86, column=11, value=df_cleaned_phase_2.at[81, 'Phase_2']  / 100)

        # workbook.save(file_path)

        # st.write(df_cleaned_phase_2,"updated one")


elif st.session_state.page == 'risk-management':
    st.markdown(
        f'<div style="background-color: {bg_color}; color: white; padding: 5px; border-radius: 50px; margin-bottom: 15px; width: 80%; text-align: center;font-size: 36px; margin-top: -50px;" class="center-text">'
        '<strong>Risk Assessment</strong>'
        '</div>', unsafe_allow_html=True)

    # Load the risk assessment data if not already in session state
    if st.session_state.risk_assessment_df is None:
        df = load_financial_model(file_path, sheet_name='Sheet1', header=4)
        df = df.iloc[:, 1:]
        df.columns = df.columns.str.strip()
        df['Mitigation cost'] = df['Mitigation cost'].fillna(0)
        st.session_state.risk_assessment_df = df
    else:
        df = st.session_state.risk_assessment_df

    # Extract risk list
    risk_list = df['Risk'].dropna().tolist()
    
    # List of editable fields
    editable_fields = [
        "Category", 
        "Base Cost Link (CAPEX/OPEX/Maintenance)", 
        "Percentage of Base Cost (%)", 
        "Recurrence (if OPEX related)", 
        "Probability of Occurrence (%)", 
        "Allocation to Government (%)", 
        "Allocation to Private Sector (%)",
        "Mitigation cost"
    ]

    # Function to ensure a risk can only be selected once
    def available_risks(index):
        selected = st.session_state.selected_risks[:index] + st.session_state.selected_risks[index+1:]
        return ["Select a risk"] + [risk for risk in risk_list if risk not in selected]

    for i in range(18):
        st.markdown(f"### Select Risk {i + 1}")

        # Dropdown to select the risk, ensure each risk can only be selected once
        selected_risk = st.selectbox(
            f"Risk {i + 1}",
            available_risks(i),
            key=f"risk_{i}",
            index=available_risks(i).index(st.session_state.selected_risks[i])
        )

        st.session_state.selected_risks[i] = selected_risk

        # If a valid risk is selected, show editable inputs
        if selected_risk != "Select a risk":
            selected_risk_data = df[df['Risk'] == selected_risk].iloc[0]

            # Use an expander to show editable fields for the selected risk
            with st.expander(f"Details for {selected_risk}"):
                for column in editable_fields:
                    # Get the current value from the DataFrame
                    current_value = selected_risk_data[column]

                    # Format the value appropriately
                    if column in ["Percentage of Base Cost (%)", "Probability of Occurrence (%)", 
                                "Allocation to Government (%)", "Allocation to Private Sector (%)"]:
                        current_value_str = f"{current_value * 100:.2f}"
                    else:
                        current_value_str = str(current_value)

                    # Display as text input for user to edit
                    new_value = st.text_input(
                        f"{column} (Risk {i + 1})", 
                        value=current_value_str,
                        key=f"text_input_{i}_{column}"
                    )

                    # Update the value in the DataFrame
                    try:
                        if column in ["Percentage of Base Cost (%)", "Probability of Occurrence (%)", 
                                    "Allocation to Government (%)", "Allocation to Private Sector (%)"]:
                            df.loc[df['Risk'] == selected_risk, column] = float(new_value) / 100
                        elif column == "Mitigation cost":
                            df.loc[df['Risk'] == selected_risk, column] = float(new_value)
                        else:
                            df.loc[df['Risk'] == selected_risk, column] = new_value
                    except ValueError:
                        pass

    # Update the session state DataFrame
    st.session_state.risk_assessment_df = df

    if st.button("Save Changes"):
        # Check if at least one risk has been selected
        if all(risk == "Select a risk" for risk in st.session_state.selected_risks):
            st.error("Please select at least one risk before saving changes.")
        else:
            try:
                file_extension = Path(file_path).suffix.lower()[1:]
                if file_extension in ['xlsx', 'xls']:
                    workbook = load_workbook(file_path)
                    sheet = workbook['Sheet1']

                # Iterate through all selected risks and update corresponding rows in the Excel sheet
                for i, selected_risk in enumerate(st.session_state.selected_risks):
                    if selected_risk != "Select a risk":
                        if selected_risk in df['Risk'].values:
                            # Find the row corresponding to the selected risk
                            risk_row = df.index[df['Risk'] == selected_risk].tolist()[0] + 6  # +5 to account for header offset in Excel

                        # Update only the changed fields in the sheet
                        for column in editable_fields:
                            col_idx = df.columns.get_loc(column) + 2  

                            if column == "Allocation to Government (%)":
                                col_idx += 1  

                            if column == "Allocation to Private Sector (%)":
                                col_idx += 1  

                            if column == "Mitigation cost":
                                col_idx += 3    

                            if column in ["Percentage of Base Cost (%)", "Probability of Occurrence (%)", 
                                        "Allocation to Government (%)", "Allocation to Private Sector (%)"]:
                                sheet.cell(row=risk_row, column=col_idx, value=df.loc[df['Risk'] == selected_risk, column].values[0])
                            elif column == "Mitigation cost":
                                sheet.cell(row=risk_row, column=col_idx, value=df.loc[df['Risk'] == selected_risk, column].values[0])
                            else:
                                sheet.cell(row=risk_row, column=col_idx, value=df.loc[df['Risk'] == selected_risk, column].values[0])

                st.success("All changes have been saved successfully!")

            except Exception as e:
                st.error(f"Error saving changes: {e}")

    # Navigation buttons
    col1, col2 = st.columns([1, 18])
    with col1:
        if st.session_state.selected_phase_option == "2":
            if check_phase2:
                st.button("Back", on_click=continue_to_phase)
            else:
                st.button("Back", on_click=continue_to_phase2)
        else:
            st.button("Back", on_click=continue_to_phase)
    with col2:
        st.button("Continue to Dashboard", on_click=continue_to_dashboard)

elif st.session_state.page == 'dashboard':

    def load_lottie_url(url: str):
        response = requests.get(url)
        if response.status_code == 200:
            return response.json()
        return None
    lottie_url1 = "https://assets10.lottiefiles.com/packages/lf20_jcikwtux.json"  
    lottie_url2 = "https://lottie.host/da208e68-3a3a-48a9-b73f-17f8925cde2a/zJ0MoRmhHJ.json"
    lottie_animation1 = load_lottie_url(lottie_url1)
    lottie_animation2 = load_lottie_url(lottie_url2)
    
    with st.sidebar:
        st.header("Dashboard Navigation")
        options = st.sidebar.radio("Select a page:", ["Data Overview","User Details","Download Report"])
        
        # Add spacer to push the back button to the bottom
        st.markdown("<br>" * 10, unsafe_allow_html=True)
        
        # Add back button with improved styling
        st.markdown(
            """
            <style>
            .sidebar-back-button {
                position: fixed;
                bottom: 20px;
                left: 20px;
                z-index: 1000;
                width: calc(100% - 40px);
                max-width: 300px;
            }
            .sidebar-back-button button {
                background-color: #000C66;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                cursor: pointer;
                font-size: 14px;
                transition: background-color 0.3s;
                width: 100%;
                text-align: center;
            }
            .sidebar-back-button button:hover {
                background-color: #000a4d;
            }
            </style>
            """,
            unsafe_allow_html=True
        )
        
        # Create a container for the back button
        st.markdown(
            '<div class="sidebar-back-button">',
            unsafe_allow_html=True
        )
        st.button("← Back to Risk Management", on_click=continue_to_risk_management, key="dashboard_back")
        st.markdown('</div>', unsafe_allow_html=True)

    if  options == "Data Overview":

        # st.write(file_path,'----------file pathhh')
        
        df = load_financial_model(file_path,sheet_name='Output',header=7)

        # st.write(df)

        if df is None or df.empty:
            st.error("Failed to load data from the Output sheet.")
        # else:
        #     st.write(df)

        # st.write(df)

        fixed_indices = [0, 1, 2,9]

        range_indices = list(range(11, 63))

        indcies_to_drop = range_indices + fixed_indices

        df_cleaned = df.drop(fixed_indices) 

        # st.write(df_cleaned)


        # Total risk values:
        df_risks = load_financial_model(file_path,sheet_name='Sheet1',header=None)

        cell_value = df_risks.iloc[43, 14]
        cell_value_2 = df_risks.iloc[43, 15]



        rounded_value_risk_total = f'{cell_value:.2f}'
        rounded_value_risk_total_2 = f'{cell_value_2:.2f}'




        df_totla_debt = load_financial_model(file_path,sheet_name='Phase_1',header=None)

        cell_value = df_totla_debt.iloc[83, 5]

        rounded_value_total_debt_1 = f'{cell_value:.2f}'

        # st.write(rounded_value_total_debt_1)


        df_totla_debt_2 = load_financial_model(file_path,sheet_name='Phase_2',header=None)

        cell_value = df_totla_debt_2.iloc[83, 5]

        rounded_value_total_debt_2 = f'{cell_value:.2f}'

        # st.write(rounded_value_total_debt_2)


        # st.write(rounded_value_risk_total)
        # st.write(rounded_value_risk_total_2)


        


        # Dashboard title
        # st.set_page_config(page_title="Financial Analytics Dashboard", layout="wide")
        st.markdown("""
            <style>
                .custom-title {
                    font-size: 32px;
                    font-weight: 800;
                    color: #FFFFFF; /* White text for contrast */
                    background-color: #000c66; 
                    padding: 15px 20px;
                    border-radius: 10px;
                    text-align: center;
                    box-shadow: 0 4px 15px rgba(0, 0, 0, 0.3); 
                    margin-bottom: 20px;
                    letter-spacing: 1px; 

                }
                
            </style>
            <div class="custom-title">Financial Analytics Dashboard</div>
            """, unsafe_allow_html=True
        )
        st.divider()

        total_unitary_charge_phase_1 = df_cleaned.loc[8, 'Phase_1']
        total_unitary_charge_phase_2 = df_cleaned.loc[8, 'Phase_2']
        equity_irr_phase_1 = df_cleaned.loc[10, 'Phase_1']
        equity_irr_phase_2 = df_cleaned.loc[10, 'Phase_2']


        equity_irr_phase_1_percentage = equity_irr_phase_1 * 100  
        equity_irr_phase_2_percentage = equity_irr_phase_2 * 100


        st.markdown("""
            <style>
                .custom-subheader {
                    color: #000c66;
                    font-size: 28px;
                    font-weight: bold;
                    margin-top: 20px; /* Optional: add space above the subheader */
                    margin-bottom: 10px; /* Optional: add space below the subheader */
                }
            </style>
            <h2 class='custom-subheader'>Key Financial Metrics</h2>
            """, unsafe_allow_html=True
        )

        cols = st.columns(8)

        def create_metric_card(col, label, value, color):
            col.markdown(
                f"""
                <div style="
                    border-radius: 10px;
                    padding: 15px;
                    text-align: center;
                    background-color: {color};
                    box-shadow: 0 4px 10px rgba(0, 0, 0, 0.1);
                    color: white;
                    margin-bottom: 15px;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    justify-content: center;
                    min-height: 124px;
                    flex-grow: 1;  /* Allow columns to grow equally */
                    margin: 10px;
                ">
                    <h4 style="margin: 0; font-size: 15px; font-weight: 800;">{label}</h4>
                    <p style="font-size: 16px; font-weight: 900; margin: 5px 0 0 0;">{value}</p>
                </div>
                """,
                unsafe_allow_html=True
            )

        col1, col2 = st.columns(2)
        create_metric_card(col1, "Total Unity Charge - Phase 1", f"{total_unitary_charge_phase_1:.2f} LE/m³", "#00C9A7")  # Navy blue
        create_metric_card(col2, "Total Unity Charge - Phase 2", f"{total_unitary_charge_phase_2:.2f} LE/m³", "#FFDD44")  # Gold/yellow

        col3, col4 = st.columns(2)
        create_metric_card(col3, "Equity IRR - Phase 1", f"{equity_irr_phase_1_percentage:.2f}%", "#17A2B8")  # Teal
        create_metric_card(col4, "Equity IRR - Phase 2", f"{equity_irr_phase_2_percentage:.2f}%", "#E74C3C")  # Red

        col5, col6 = st.columns(2)
        create_metric_card(col5, "Cost of risk on the Government (LE or $)", f"{rounded_value_risk_total}", "#E67E22")  
        create_metric_card(col6, "Cost of risk on the Private (LE or $)", f"{rounded_value_risk_total_2}", "#999B27")


        col7, col8 = st.columns(2)
        create_metric_card(col5, "Total Debt-Phase 1", f"{rounded_value_total_debt_1}", "#1E90FF")  
        create_metric_card(col6, "Total Debt- Phase 2", f"{rounded_value_total_debt_2}", "#FF7F50")

        st.divider()



        tariff_labels = df_cleaned['Unnamed: 4'].iloc[:5]  

        # st.write(df_cleaned)

        # st.write(tariff_labels)
        tariff_values_phase_1 = df_cleaned['Phase_1'].iloc[:5].round(2)  
        tariff_values_phase_2 = df_cleaned['Phase_2'].iloc[:5].round(2)  

        fig_pie_phase_1 = px.pie(
            names=tariff_labels,  
            values=tariff_values_phase_1,  
            title="Unitary Charge Breakdown - Phase 1",
            color_discrete_sequence=["#00C9A7", "#FFDD44", "#17A2B8", "#E74C3C", "#E67E22"],  
            hole=0.4  
        )

        fig_pie_phase_2 = px.pie(
            names=tariff_labels, 
            values=tariff_values_phase_2,  
            title="Unitary Charge Breakdown - Phase 2",
            color_discrete_sequence=["#00C9A7", "#FFDD44", "#17A2B8", "#E74C3C", "#E67E22"],  
            hole=0.4 
        )

        # Update traces for better text positioning
        fig_pie_phase_1.update_traces(
            textposition='auto',  # Automatically decide based on space
            textinfo='value+percent',  # Show both value and percent
            insidetextorientation='radial',  # Make the text readable inside
            marker=dict(line=dict(color='white', width=1)),  # Add borders for clarity
            # pull=[0, 0, 0, 0.1, 0.2],  # Pull specific small slices for better clarity
            textfont=dict(size=14)  # Adjust font size for better readability
        )

        fig_pie_phase_2.update_traces(
            textposition='auto', 
            textinfo='value+percent',  
            insidetextorientation='radial', 
            marker=dict(line=dict(color='white', width=1)), 
            # pull=[0, 0, 0, 0.1, 0.2],
            textfont=dict(size=14)
        )

        # Update layout to match better spacing
        for fig in [fig_pie_phase_1, fig_pie_phase_2]:
            fig.update_layout(
                title_font=dict(size=18, color='darkblue'),
                showlegend=True,
                legend=dict(orientation="h", y=-0.2, title=None),  # Legend below the chart
                margin=dict(t=40, b=20, l=0, r=0)  # Add margins
            )

        # Display the charts side by side in Streamlit
        cols_pie = st.columns(2)
        with cols_pie[0]:
            st.plotly_chart(fig_pie_phase_1, use_container_width=True)
        with cols_pie[1]:
            st.plotly_chart(fig_pie_phase_2, use_container_width=True)


    elif options == "User Details":
        col1,col2 = st.columns([3,5])
        with col1:
            st.markdown("<br><br>", unsafe_allow_html=True)
            # st_lottie(lottie_animation1, height=350, key="home1")
            if lottie_animation2 is not None:
                st_lottie(lottie_animation2, height=400, key="home2")
        with col2:
            st.markdown(
            f'<br><div style="background-color: #3f3f3f; color: white; padding: 10px; border-radius: 50px; width:70%; margin-bottom: 15px; font-size: 28px;text-align: center;">'
            '<strong>USER DETAILS</strong>'
            '</div>', unsafe_allow_html=True)
            custom_css = """
            <style>
            .custom-text {
                font-family: 'Arial', sans-serif; /* Change the font family here */
                font-size: 24px; /* Change the text size here */
                color: #333; /* Change the text color here */
                padding: 10px; /* Add some padding if needed */
                background-color: #FFD700; /* Change the background color here */
                border-radius: 50px; /* Optional: Add rounded corners */
                text-align: center;
                margin-bottom: 15px;
                width:70%;
            }
            </style>
            """

            st.markdown(custom_css, unsafe_allow_html=True)

            st.markdown(f'<div class="custom-text">Username: {st.session_state.field1}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="custom-text">Project: {st.session_state.field2}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="custom-text">Capacity: {st.session_state.field3}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="custom-text">Location: {st.session_state.field4}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="custom-text">Date: {st.session_state.field5}</div>', unsafe_allow_html=True)
    elif options == "Download Report":
        if lottie_animation1 is not None:
            st_lottie(lottie_animation1, height=400, key="home3")
        def generate_report():
            data = {
                "Field": ["Username", "Project", "Capacity", "Location", "Option"],
                "Value": [
                    st.session_state.field1,
                    st.session_state.field2,
                    st.session_state.field3,
                    st.session_state.field4,
                    st.session_state.field5
                ]
            }
            df = pd.DataFrame(data)
            
            buffer = StringIO()
            df.to_csv(buffer, index=False)
            return buffer.getvalue()

        # Download button
        csv_report = generate_report()
        st.markdown("""
            <style>
            .download-button {
                display: flex;
                justify-content: center;
                align-items: center;
                
            }
            .download-button a {
                display: inline-block;
                padding: 20px 40px;
                font-size: 18px;
                color: white;
                background-color: #007bff;
                border-radius: 5px;
                text-decoration: none;
                text-align: center;
            }
            .download-button a:hover {
                background-color: #0056b3;
            }
            </style>
            <div class="download-button">
                <a href="data:text/csv;charset=utf-8,{csv_report}" download="report.csv">Download Report</a>
            </div>
        """, unsafe_allow_html=True)
        # st.download_button(
        #     label="Download Report",
        #     data=csv_report,
        #     file_name="report.csv",
        #     mime="text/csv"
        # )
